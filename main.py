import os
import logging
import datetime
import tempfile
import subprocess

import whisper
from openpyxl import Workbook, load_workbook
from telegram import Update, InputFile
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters
)

TOKEN = "8407369465:AAFJ8MCRIkWoO2HiETILry7XeuHf81T1DBw"
EXCEL_FILE = "expenses.xlsx"
WHISPER_MODEL_NAME = os.getenv("WHISPER_MODEL", "tiny")  # tiny/small

logging.basicConfig(level=logging.INFO)

os.environ["CUDA_VISIBLE_DEVICES"] = ""
os.environ["PYTORCH_ENABLE_MPS_FALLBACK"] = "1"

model = whisper.load_model(WHISPER_MODEL_NAME)
WHISPER_ARGS = {"language": "ar", "fp16": False}

if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    ws.append(["التاريخ", "الوصف", "المبلغ", "التصنيف", "النوع"])
    wb.save(EXCEL_FILE)

def parse_text(text: str):
    import re
    amount = ""
    m = re.search(r"(\d{1,6})\s*ريال", text)
    if m:
        amount = m.group(1)

    t = text
    if any(k in t for k in ["بنزين", "وقود", "محطة", "سيارة"]):
        category = "السيارة"
    elif any(k in t for k in ["ملابس", "تيشيرت", "بنطلون", "عباية"]):
        category = "الملابس"
    elif any(k in t for k in ["مطعم", "عشاء", "غداء", "فطور", "قهوة", "مشروب", "بيتزا", "برغر"]):
        category = "مطاعم/قهوة"
    elif any(k in t for k in ["دخل", "حوّلت", "حوّل", "ربح", "مكسب", "وصلني"]):
        category = "دخل"
    else:
        category = "غير مصنف"

    tx_type = "ربح" if category == "دخل" or any(k in t for k in ["دخل", "ربح", "مكسب"]) else "خسارة"
    return amount, category, tx_type


def save_row(desc: str, amount: str, category: str, tx_type: str):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.append([now, desc, amount, category, tx_type])
    wb.save(EXCEL_FILE)


def ffmpeg_to_wav(src_path: str, dst_wav: str):
    # 16k mono wav
    subprocess.run(
        ["ffmpeg", "-y", "-i", src_path, "-ar", "16000", "-ac", "1", "-vn", dst_wav],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, check=True
    )


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "🎙️ أرسل **مقطع صوتي أو ڤويس** بكلام عام مثل:\n"
        "• دفعت 40 ريال بنزين\n"
        "• قهوة 12 ريال\n"
        "• جاني دخل 100 ريال\n"
        "وسأسجّله في الإكسل.\n\n"
        "📄 للأرشيف: /report",
        parse_mode="Markdown"
    )


async def report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not os.path.exists(EXCEL_FILE):
        await update.message.reply_text("لا يوجد سجل حتى الآن.")
        return
    await update.message.reply_document(InputFile(EXCEL_FILE))


async def handle_media(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        file_id = None
        if update.message.voice:
            file_id = update.message.voice.file_id
        elif update.message.audio:
            file_id = update.message.audio.file_id
        elif update.message.document and str(update.message.document.mime_type).startswith("audio/"):
            file_id = update.message.document.file_id

        if not file_id:
            return

        tg_file = await context.bot.get_file(file_id)

        with tempfile.NamedTemporaryFile(delete=False, suffix=".bin") as src:
            await tg_file.download_to_drive(src.name)
            src_path = src.name

        with tempfile.NamedTemporaryFile(delete=False, suffix=".wav") as dst:
            wav_path = dst.name

        ffmpeg_to_wav(src_path, wav_path)

        result = model.transcribe(wav_path, **WHISPER_ARGS)
        text = (result.get("text") or "").strip()

        for p in (src_path, wav_path):
            try:
                os.remove(p)
            except Exception:
                pass

        if not text:
            await update.message.reply_text("❌ لم أفهم التسجيل. اذكر المبلغ بصيغة «XX ريال».")
            return

        amount, category, tx_type = parse_text(text)
        if not amount:
            await update.message.reply_text(
                f"🗒️ النص المستخرج:\n{text}\n\n"
                "❌ لم أجد مبلغًا بصيغة «XX ريال». أعد المحاولة واذكر المبلغ."
            )
            return

        save_row(text, amount, category, tx_type)
        await update.message.reply_text(
            f"✅ تم التسجيل:\n"
            f"• الوصف: {text}\n"
            f"• المبلغ: {amount} ريال\n"
            f"• التصنيف: {category}\n"
            f"• النوع: {tx_type}"
        )

    except FileNotFoundError:
        await update.message.reply_text("❌ ffmpeg غير متوفر. أعِد النشر وتأكد من وجوده في render.yaml.")
        logging.exception("ffmpeg missing")
    except Exception:
        await update.message.reply_text("❌ حدث خطأ أثناء معالجة الصوت.")
        logging.exception("media handler error")


def run():
    app = ApplicationBuilder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("report", report))
    app.add_handler(MessageHandler(filters.VOICE | filters.AUDIO | filters.Document.AUDIO, handle_media))
    app.run_polling(close_loop=False)


if __name__ == "__main__":
    run()
